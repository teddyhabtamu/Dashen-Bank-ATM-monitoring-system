"""Import the 1,202 real ATMs from Abinet's Excel into atm_locations.

Usage:
    python3 scripts/import_atms_from_excel.py                    # preview only
    python3 scripts/import_atms_from_excel.py --apply             # write changes
    python3 scripts/import_atms_from_excel.py --apply --csv-only  # only rewrite the CSV, skip DB

Normalises vendors (Moti engineering/NCR -> NCR, tech24et/GRG -> GRG),
normalises district names to uppercase, merges with existing ATMs by
terminal_id, and generates atm_id values for new entries.
"""

import argparse
import csv
import re
import sys

EXCEL_PATH = "docs/ATM INFORMATION.xlsx"
CSV_PATH = "config/postgres/atm_locations.csv"

# Normalise these vendor strings to our internal codes.
VENDOR_MAP = {
    "Moti engineering/NCR": "NCR",
    "tech24et/GRG": "GRG",
}

# Normalise district names (source: consistent uppercase).
# Some entries appear with mixed case; normalise to uppercase.
DISTRICT_NORMALISE = {
    "Adama": "ADAMA",
    "Bahir Dar": "BAHIR DAR",
    "Dessie": "DESSIE",
    "Dire Dawa": "DIRE DAWA",
    "JImma": "JIMMA",
    "Jimma": "JIMMA",
    "Mekelle": "MEKELLE",
    "Hawasa": "HAWASA",
    "Nekemte": "NEKEMTE",
    "Wolaita": "WOLAITA",
    "South West": "SOUTH WEST",
}


def normalise_district(name):
    if not name or name == "None":
        return None
    name = name.strip()
    upper = name.upper()
    if name in DISTRICT_NORMALISE:
        return DISTRICT_NORMALISE[name]
    return upper


def load_excel():
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    atms = []
    for r in range(2, ws.max_row + 1):
        tid = str(ws.cell(r, 2).value or "").strip()
        name = str(ws.cell(r, 3).value or "").strip()
        ip = str(ws.cell(r, 4).value or "").strip()
        branch = str(ws.cell(r, 5).value or "").strip()
        district = str(ws.cell(r, 6).value or "").strip()
        vendor_raw = str(ws.cell(r, 7).value or "").strip()

        if not tid:
            continue  # skip blank rows

        vendor = VENDOR_MAP.get(vendor_raw, vendor_raw)
        district_norm = normalise_district(district)

        atms.append({
            "terminal_id": tid,
            "atm_name": name,
            "ip_address": ip if ip and ip != "None" else None,
            "branch": branch if branch and branch != "None" else None,
            "district": district_norm,
            "vendor": vendor,
        })
    return atms


def load_existing_csv():
    existing = {}
    try:
        with open(CSV_PATH) as f:
            reader = csv.DictReader(f)
            for row in reader:
                tid = row.get("terminal_id", "").strip()
                if tid:
                    existing[tid] = row
    except FileNotFoundError:
        pass
    return existing


def next_atm_id(existing, new_atms):
    """Generate ATM-XXX IDs avoiding collisions."""
    used = set()
    for row in existing.values():
        aid = row.get("atm_id", "").strip()
        if aid:
            used.add(aid)
    # Also check reserved prefixes
    prefix = "ATM-"
    max_num = 0
    for aid in used:
        if aid.startswith(prefix):
            try:
                n = int(aid[len(prefix):])
                if n > max_num:
                    max_num = n
            except ValueError:
                pass
    next_num = max_num + 1
    for a in new_atms:
        a["atm_id"] = f"{prefix}{next_num:03d}"
        next_num += 1


def main():
    parser = argparse.ArgumentParser(description="Import ATMs from Excel")
    parser.add_argument("--apply", action="store_true", help="Actually write changes")
    parser.add_argument("--csv-only", action="store_true", help="Only update CSV, skip DB")
    args = parser.parse_args()

    print("Loading Excel...")
    excel_atms = load_excel()
    print(f"  {len(excel_atms)} ATMs in Excel")

    # Separate new vs existing by terminal_id
    existing = load_existing_csv()
    print(f"  {len(existing)} existing ATMs in CSV")

    new_atms = []
    skip_count = 0
    for a in excel_atms:
        if a["terminal_id"] in existing:
            skip_count += 1
        else:
            new_atms.append(a)

    print(f"  {skip_count} already exist, {len(new_atms)} new")

    # Generate atm_id for new entries
    next_atm_id(existing, new_atms)

    # Preview
    if new_atms:
        print("\n  First 5 new ATMs:")
        for a in new_atms[:5]:
            print(f"    {a['atm_id']} | {a['terminal_id']} | {a['atm_name'][:25]:25s} | "
                  f"{a['vendor']:3s} | {a['district'] or '':10s} | {a['ip_address'] or '':15s}")
        if len(new_atms) > 5:
            print(f"    ... and {len(new_atms) - 5} more")

    # Vendor check
    vendors = {}
    for a in excel_atms:
        vendors[a["vendor"]] = vendors.get(a["vendor"], 0) + 1
    print(f"\n  Vendor breakdown: {vendors}")

    # District breakdown
    districts = {}
    for a in excel_atms:
        d = a["district"] or "UNKNOWN"
        districts[d] = districts.get(d, 0) + 1
    print(f"  Districts: {len(districts)}")
    for d, c in sorted(districts.items(), key=lambda x: -x[1])[:10]:
        print(f"    {d}: {c}")

    if not args.apply:
        print("\n  DRY RUN — no changes made. Re-run with --apply to write.")
        return

    # Rewrite CSV merged
    fieldnames = ["atm_id", "branch", "district", "city", "region", "latitude",
                  "longitude", "terminal_id", "vendor", "model", "install_date",
                  "status", "ip_address", "atm_name"]

    # Build a lookup for Excel data (to enrich existing CSV rows)
    excel_lookup = {a["terminal_id"]: a for a in excel_atms}

    all_rows = list(existing.values())
    for row in all_rows:
        tid = row.get("terminal_id", "").strip()
        if tid in excel_lookup:
            ea = excel_lookup[tid]
            if not row.get("atm_name", "").strip():
                row["atm_name"] = ea.get("atm_name", "") or ""
            if not row.get("ip_address", "").strip() and ea.get("ip_address"):
                row["ip_address"] = ea["ip_address"]

    for a in new_atms:
        all_rows.append({
            "atm_id": a["atm_id"],
            "branch": a["branch"] or "",
            "district": a["district"] or "",
            "city": "",
            "region": "",
            "latitude": "",
            "longitude": "",
            "terminal_id": a["terminal_id"],
            "vendor": a["vendor"],
            "model": "",
            "install_date": "",
            "status": "active",
            "ip_address": a["ip_address"] or "",
            "atm_name": a["atm_name"] or "",
        })

    with open(CSV_PATH, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_rows)
    print(f"\n  CSV written: {CSV_PATH} ({len(all_rows)} rows)")

    if args.csv_only:
        print("  CSV-only mode — DB not updated.")
        return

    # Update DB via SQL
    import psycopg2
    from db import get_db

    conn = get_db()
    cur = conn.cursor()

    # Ensure ip_address column
    cur.execute("ALTER TABLE atm_locations ADD COLUMN IF NOT EXISTS ip_address VARCHAR(45)")

    # Upsert new ATMs
    inserted = 0
    for a in new_atms:
        cur.execute("""
            INSERT INTO atm_locations
                (atm_id, terminal_id, branch, district, vendor, atm_name, ip_address, status)
            VALUES (%s, %s, %s, %s, %s, %s, %s, 'active')
            ON CONFLICT (atm_id) DO NOTHING
        """, (a["atm_id"], a["terminal_id"], a["branch"], a["district"],
              a["vendor"], a["atm_name"], a["ip_address"]))
        if cur.rowcount:
            inserted += 1

    # Update ip_address for existing ATMs where we now have it
    updated = 0
    for a in excel_atms:
        if a["terminal_id"] in existing and a["ip_address"]:
            cur.execute("""
                UPDATE atm_locations SET ip_address = %s
                WHERE terminal_id = %s AND ip_address IS NULL
            """, (a["ip_address"], a["terminal_id"]))
            if cur.rowcount:
                updated += 1

    conn.commit()
    cur.close()
    conn.close()

    print(f"  DB inserted: {inserted} new, updated IPs: {updated}")
    print("  Done!")


if __name__ == "__main__":
    main()

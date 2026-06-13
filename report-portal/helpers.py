from flask import send_file
import openpyxl, csv, io, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle
from datetime import datetime
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                Paragraph, Spacer, HRFlowable, PageBreak, Image)
from reportlab.lib.enums import TA_CENTER

LOGO_PATH = os.path.join(os.path.dirname(__file__), 'static', 'logo.png')

DASHEN_BLUE  = '273274'
DASHEN_GOLD  = 'FDD79A'
ALT_ROW      = 'F8FAFC'
BORDER_COLOR = 'E2E8F0'


# ─── EXCEL HELPERS ─────────────────────────────────────────────────────────────

def xl_header(ws, title, days, atm):
    """Write the 3-row branded header block."""
    ws.merge_cells('A1:H1')
    c = ws['A1']
    c.value = 'DASHEN BANK S.C. — ATM MONITORING SYSTEM'
    c.font = Font(bold=True, size=13, color=DASHEN_BLUE)
    c.fill = PatternFill('solid', fgColor='FFFDE7')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    ws.merge_cells('A2:H2')
    c = ws['A2']
    c.value = title
    c.font = Font(bold=True, size=11, color=DASHEN_BLUE)
    c.fill = PatternFill('solid', fgColor='FFFDE7')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    ws.merge_cells('A3:H3')
    c = ws['A3']
    c.value = (
        f'Period: Last {days} days  |  '
        f'ATM: {"All ATMs" if atm == "all" else atm}  |  '
        f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} EAT'
    )
    c.font = Font(size=9, color='64748B', italic=True)
    c.fill = PatternFill('solid', fgColor='FFFDE7')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 14
    ws.append([])  # blank spacer row


def xl_style_row(ws, row, ncols):
    """Style the column-header row: dark blue background, white bold text."""
    thin = Side(style='thin', color=BORDER_COLOR)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.fill = PatternFill('solid', fgColor=DASHEN_BLUE)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[row].height = 20


def xl_style_data_rows(ws, start_row, ncols):
    """Apply alternating row colours and borders to data rows."""
    thin = Side(style='thin', color=BORDER_COLOR)
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=ncols):
        is_alt = (row[0].row - start_row) % 2 == 1
        bg = ALT_ROW if is_alt else 'FFFFFF'
        for cell in row:
            cell.fill = PatternFill('solid', fgColor=bg)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.font = Font(size=9)


def xl_autosize(ws):
    for col in ws.columns:
        mx = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 4, 52)


def xl_send(wb, name):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Dashen_ATM_{name}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


def xl_build_sheet(wb, sheet_name, title, days, atm, headers, rows):
    """Utility: create a fully styled sheet inside an existing workbook."""
    ws = wb.create_sheet(sheet_name)
    xl_header(ws, title, days, atm)
    ws.append(headers)
    xl_style_row(ws, ws.max_row, len(headers))
    data_start = ws.max_row + 1
    for r in rows:
        ws.append([str(v) if v is not None else '—' for v in r])
    xl_style_data_rows(ws, data_start, len(headers))
    xl_autosize(ws)
    return ws


def xl_append_section(ws, section_title, headers, rows):
    """Utility: append a styled section to an existing sheet."""
    if ws.max_row > 5:  # If there is already data (past the header)
        ws.append([])
        ws.append([])

    title_row = ws.max_row + 1
    ws.append([section_title])
    c = ws.cell(row=title_row, column=1)
    c.font = Font(bold=True, size=11, color=DASHEN_BLUE)

    ws.append(headers)
    hdr_row = ws.max_row
    xl_style_row(ws, hdr_row, len(headers))
    
    data_start = ws.max_row + 1
    for r in rows:
        ws.append([str(v) if v is not None else '—' for v in r])
    xl_style_data_rows(ws, data_start, len(headers))



# ─── CSV HELPERS ───────────────────────────────────────────────────────────────

def csv_send(headers, rows, name, title='', days=0, atm='all'):
    """Write a CSV with a branded metadata block at the top."""
    buf = io.StringIO()
    w = csv.writer(buf)
    # Branded header block
    w.writerow(['DASHEN BANK S.C. — ATM MONITORING SYSTEM'])
    if title:
        w.writerow([title])
    w.writerow([
        f'Period: Last {days} days'
        if days else '',
        f'ATM: {"All ATMs" if atm == "all" else atm}',
        f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} EAT'
    ])
    w.writerow([])  # blank spacer
    w.writerow(headers)
    w.writerows([[str(v) if v is not None else '—' for v in r] for r in rows])
    w.writerow([])
    w.writerow(['Dashen Bank ATM Monitoring System | Confidential'])
    buf.seek(0)
    return send_file(
        io.BytesIO(buf.getvalue().encode('utf-8-sig')),  # utf-8-sig for Excel compatibility
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'Dashen_ATM_{name}_{datetime.now().strftime("%Y%m%d")}.csv'
    )


# ─── PDF HELPERS ───────────────────────────────────────────────────────────────

def _pdf_kpi_block(story, kpis):
    """Append a KPI summary card section to the story."""
    P = lambda txt, **kw: Paragraph(txt, ParagraphStyle('kpi', **kw))

    story.append(Spacer(1, 0.5 * cm))
    story.append(P('<font color="#0F2557" size="13"><b>Performance Summary</b></font>',
                    spaceBefore=4, spaceAfter=12))

    ncols = min(len(kpis), 4)
    page_w = landscape(A4)[0] - 3 * cm
    cw = page_w / ncols

    labels_row = []
    values_row = []
    for label, value in kpis:
        labels_row.append(
            P(f'<font size="8" color="#64748B">{label}</font>', alignment=TA_CENTER))
        values_row.append(
            P(f'<font size="20" color="#0F2557"><b>{value}</b></font>', alignment=TA_CENTER))

    while len(labels_row) < ncols:
        labels_row.append(P(''))
        values_row.append(P(''))

    t = Table([labels_row, values_row], colWidths=[cw] * ncols)
    accent_colors = ['#0F2557', '#059669', '#D97706', '#DC2626']
    style_cmds = [
        ('ALIGN',       (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME',    (0, 0), (-1,  0), 'Helvetica'),
        ('FONTNAME',    (0, 1), (-1,  1), 'Helvetica-Bold'),
        ('GRID',        (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('BACKGROUND',  (0, 0), (-1, -1), colors.white),
        ('TOPPADDING',  (0, 0), (-1,  0), 14),
        ('BOTTOMPADDING', (0, 0), (-1,  0), 6),
        ('TOPPADDING',  (0, 1), (-1,  1), 6),
        ('BOTTOMPADDING', (0, 1), (-1,  1), 14),
    ]
    for i in range(ncols):
        style_cmds.append(
            ('LINEABOVE', (i, 0), (i, 1), 3, colors.HexColor(accent_colors[i % 4])))
    t.setStyle(TableStyle(style_cmds))
    story.append(t)
    story.append(Spacer(1, 1 * cm))


def _pdf_header_block(story, title, days, atm):
    """Shared PDF header: logo + title + gold rule."""
    P = lambda txt, **kw: Paragraph(txt, ParagraphStyle('_h', **kw))
    if os.path.exists(LOGO_PATH):
        img = Image(LOGO_PATH, width=5 * cm, height=1.5 * cm, kind='proportional')
        img.hAlign = 'CENTER'
        story.append(img)
        story.append(Spacer(1, 0.3 * cm))
    story += [
        P('<font color="#0F2557" size="14"><b>DASHEN BANK S.C. — ATM MONITORING SYSTEM</b></font>',
          alignment=TA_CENTER, spaceAfter=4),
        P(f'<font color="#0F2557" size="11"><b>{title}</b></font>',
          alignment=TA_CENTER, spaceAfter=4),
        P(f'<font color="#64748B" size="9">Period: Last {days} days  |  '
          f'ATM: {"All ATMs" if atm == "all" else atm}  |  '
          f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} EAT</font>',
          alignment=TA_CENTER, spaceAfter=2),
        HRFlowable(width='100%', thickness=2, color=colors.HexColor('#C9A84C'), spaceAfter=12),
    ]


def pdf_send(title, headers, rows, days, atm, name, kpis=None):
    """Generate a single-section branded PDF report with optional KPI summary page."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            rightMargin=1.5 * cm, leftMargin=1.5 * cm,
                            topMargin=2 * cm, bottomMargin=1.5 * cm)
    P = lambda txt, **kw: Paragraph(txt, ParagraphStyle('_p', **kw))

    story = []
    _pdf_header_block(story, title, days, atm)

    if kpis:
        _pdf_kpi_block(story, kpis)
        story.append(PageBreak())
        # Re-use same header on the data page
        _pdf_header_block(story, title, days, atm)

    story.append(mktable(headers, rows))
    story.append(Spacer(1, 0.5 * cm))
    story.append(P(
        f'<font size="8" color="#64748B">Dashen Bank ATM Monitoring System  |  Confidential  |  '
        f'{datetime.now().strftime("%Y-%m-%d")}</font>',
        alignment=TA_CENTER
    ))
    doc.build(story)
    buf.seek(0)
    return send_file(buf, mimetype='application/pdf', as_attachment=True,
                     download_name=f'Dashen_ATM_{name}_{datetime.now().strftime("%Y%m%d")}.pdf')


def mktable(headers, rows):
    """Build a fully styled reportlab Table for use in PDFs."""
    data = [headers] + [[str(v) if v is not None else '—' for v in r] for r in rows]
    page_w = landscape(A4)[0] - 3 * cm
    cw = page_w / len(headers)
    t = Table(data, colWidths=[cw] * len(headers), repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1,  0), colors.HexColor('#0F2557')),
        ('TEXTCOLOR',     (0, 0), (-1,  0), colors.white),
        ('FONTNAME',      (0, 0), (-1,  0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1,  0), 9),
        ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('FONTSIZE',      (0, 1), (-1, -1), 8),
        ('GRID',          (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('ROWHEIGHT',     (0, 0), (-1, -1), 20),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    return t

from flask import Blueprint, request, jsonify, send_file
from blueprints.auth import login_required, role_required, ROLE_VIEWER, ROLE_OPERATOR, ROLE_ADMIN
from audit import log_action
import openpyxl, io, os, csv
from collections import Counter
from db import get_db
from helpers import (xl_header, xl_style_row, xl_style_data_rows, xl_autosize,
                     xl_send, xl_build_sheet, xl_append_section, csv_send, pdf_send, mktable,
                     _pdf_header_block, LOGO_PATH)
from datetime import datetime
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                Paragraph, Spacer, HRFlowable, PageBreak, Image)
from reportlab.lib.enums import TA_CENTER

bp = Blueprint('report', __name__)


@bp.route('/api/atms')
@login_required
@role_required(ROLE_VIEWER, ROLE_OPERATOR, ROLE_ADMIN)
def api_atms():
    try:
        with get_db() as conn:
            cur = conn.cursor()
            cur.execute("SELECT atm_id, branch FROM atm_locations ORDER BY atm_id")
            rows = cur.fetchall(); cur.close()
        return jsonify([{'id': r[0], 'branch': r[1]} for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/api/stats')
@login_required
@role_required(ROLE_VIEWER, ROLE_OPERATOR, ROLE_ADMIN)
def api_stats():
    try:
        with get_db() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT COUNT(*),
                    ROUND(100.0*SUM(CASE WHEN status='APPROVED' THEN 1 ELSE 0 END)/NULLIF(COUNT(*),0),1),
                    COALESCE(SUM(CASE WHEN status='APPROVED' AND txn_type='WITHDRAWAL'
                        AND recorded_at>=CURRENT_DATE THEN amount ELSE 0 END),0),
                    (SELECT COUNT(*) FROM atm_locations)
                FROM atm_transactions
                WHERE recorded_at >= NOW() - INTERVAL '7 days'
            """)
            r = cur.fetchone(); cur.close()
        return jsonify({'total_txns': int(r[0]), 'success_rate': float(r[1]), 'cash_today': float(r[2]), 'atm_count': int(r[3])})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ─── TRANSACTION SUMMARY ────────────────────────────────────────────────────────

@bp.route('/report/transaction/<fmt>')
@login_required
@role_required(ROLE_VIEWER, ROLE_OPERATOR, ROLE_ADMIN)
def report_transaction(fmt):
    days = int(request.args.get('days', 7))
    atm  = request.args.get('atm', 'all')

    with get_db() as conn:
        cur = conn.cursor()
        sql = """
            SELECT
                atm_id                                                        AS "ATM",
                branch                                                        AS "Branch",
                COUNT(*)                                                      AS "Total Txns",
                SUM(CASE WHEN status='APPROVED' THEN 1 ELSE 0 END)           AS "Approved",
                SUM(CASE WHEN status='DECLINED' THEN 1 ELSE 0 END)           AS "Declined",
                SUM(CASE WHEN status='ERROR'    THEN 1 ELSE 0 END)           AS "Errors",
                ROUND(100.0*SUM(CASE WHEN status='APPROVED' THEN 1 ELSE 0 END)
                    /NULLIF(COUNT(*),0),2)                                    AS "Success Rate",
                COALESCE(SUM(CASE WHEN status='APPROVED' AND txn_type='WITHDRAWAL'
                    THEN amount ELSE 0 END),0)                               AS "Cash Dispensed ETB"
            FROM atm_transactions
            WHERE recorded_at >= NOW() - INTERVAL %s
        """
        params = [f"{days} days"]
        if atm != 'all':
            sql += " AND atm_id = %s"
            params.append(atm)
        sql += " GROUP BY atm_id, branch ORDER BY atm_id"

        cur.execute(sql, params)
        rows = cur.fetchall(); headers = [d[0] for d in cur.description]
        cur.close()

    title = 'Transaction Summary Report'
    if fmt == 'excel':
        wb = openpyxl.Workbook(); wb.remove(wb.active)
        xl_build_sheet(wb, 'Transactions', title, days, atm, headers, rows)
        log_action('EXPORT', f'Transaction report (excel) - days={days}, atm={atm}')
        return xl_send(wb, 'Transactions')
    elif fmt == 'pdf':
        total = sum(r[2] for r in rows)
        approved = sum(r[3] for r in rows)
        declined = sum(r[4] for r in rows)
        errors_cnt = sum(r[5] for r in rows)
        cash = sum(r[7] for r in rows)
        rate = round(approved / total * 100, 1) if total else 0
        kpis = [
            ('Total Transactions', f'{total:,}'),
            ('Success Rate', f'{rate}%'),
            ('Declined / Errors', f'{declined + errors_cnt:,}'),
            ('Cash Dispensed', f'ETB {cash:,.0f}'),
        ]
        log_action('EXPORT', f'Transaction report (pdf) - days={days}, atm={atm}')
        return pdf_send(title, headers, rows, days, atm, 'Transactions', kpis=kpis)
    else:
        log_action('EXPORT', f'Transaction report (csv) - days={days}, atm={atm}')
        return csv_send(headers, rows, 'Transactions', title=title, days=days, atm=atm)


# ─── CASH LEVEL ─────────────────────────────────────────────────────────────────

@bp.route('/report/cash/<fmt>')
@login_required
@role_required(ROLE_VIEWER, ROLE_OPERATOR, ROLE_ADMIN)
def report_cash(fmt):
    days = int(request.args.get('days', 7))
    atm  = request.args.get('atm', 'all')

    with get_db() as conn:
        cur = conn.cursor()
        sql = """
            SELECT
                atm_id                             AS "ATM",
                branch                             AS "Branch",
                COALESCE(SUM(amount),0)            AS "Total Dispensed ETB",
                COALESCE(ROUND(AVG(amount),2),0)   AS "Avg Withdrawal ETB",
                COUNT(*)                           AS "Withdrawal Count",
                COALESCE(MAX(amount),0)            AS "Largest Withdrawal ETB"
            FROM atm_transactions
            WHERE status='APPROVED' AND txn_type='WITHDRAWAL'
            AND recorded_at >= NOW() - INTERVAL %s
        """
        params = [f"{days} days"]
        if atm != 'all':
            sql += " AND atm_id = %s"
            params.append(atm)
        sql += " GROUP BY atm_id, branch ORDER BY 3 DESC"

        cur.execute(sql, params)
        rows = cur.fetchall(); headers = [d[0] for d in cur.description]
        cur.close()

    title = 'Cash Level & Dispensing Report'
    if fmt == 'excel':
        wb = openpyxl.Workbook(); wb.remove(wb.active)
        xl_build_sheet(wb, 'Cash', title, days, atm, headers, rows)
        log_action('EXPORT', f'Cash report (excel) - days={days}, atm={atm}')
        return xl_send(wb, 'Cash')
    elif fmt == 'pdf':
        total_dispensed = sum(r[2] for r in rows)
        total_withdrawals = sum(r[4] for r in rows)
        avg_withdrawal = round(total_dispensed / total_withdrawals, 0) if total_withdrawals else 0
        largest = max((r[5] for r in rows), default=0)
        kpis = [
            ('Total Dispensed', f'ETB {total_dispensed:,.0f}'),
            ('Total Withdrawals', f'{total_withdrawals:,}'),
            ('Avg Withdrawal', f'ETB {avg_withdrawal:,.0f}'),
            ('Largest Withdrawal', f'ETB {largest:,.0f}'),
        ]
        log_action('EXPORT', f'Cash report (pdf) - days={days}, atm={atm}')
        return pdf_send(title, headers, rows, days, atm, 'Cash', kpis=kpis)
    else:
        log_action('EXPORT', f'Cash report (csv) - days={days}, atm={atm}')
        return csv_send(headers, rows, 'Cash', title=title, days=days, atm=atm)


ERROR_DESCRIPTIONS = {
    '3A7F': 'Cash Jam',
    'B2C1': 'Card Read Error',
    '44AA': 'Network Timeout',
}

def _describe_error(code):
    return ERROR_DESCRIPTIONS.get(code, 'Unknown')

# ─── ERROR & INCIDENT ────────────────────────────────────────────────────────────

@bp.route('/report/error/<fmt>')
@login_required
@role_required(ROLE_VIEWER, ROLE_OPERATOR, ROLE_ADMIN)
def report_error(fmt):
    days = int(request.args.get('days', 7))
    atm  = request.args.get('atm', 'all')

    with get_db() as conn:
        cur = conn.cursor()
        sql = """
            SELECT
                atm_id      AS "ATM",
                branch      AS "Branch",
                error_code  AS "Error Code",
                COUNT(*)    AS "Occurrences",
                TO_CHAR(MIN(recorded_at) AT TIME ZONE 'Africa/Addis_Ababa','YYYY-MM-DD HH24:MI') AS "First Seen",
                TO_CHAR(MAX(recorded_at) AT TIME ZONE 'Africa/Addis_Ababa','YYYY-MM-DD HH24:MI') AS "Last Seen"
            FROM atm_transactions
            WHERE status='ERROR' AND error_code IS NOT NULL
            AND recorded_at >= NOW() - INTERVAL %s
        """
        params = [f"{days} days"]
        if atm != 'all':
            sql += " AND atm_id = %s"
            params.append(atm)
        sql += " GROUP BY atm_id, branch, error_code ORDER BY 4 DESC"

        cur.execute(sql, params)
        rows = cur.fetchall(); headers = [d[0] for d in cur.description]
        cur.close()

    # Augment rows with error-code description
    desc_ndx = headers.index('Error Code') + 1
    headers.insert(desc_ndx, 'Description')
    rows = [list(r) for r in rows]
    for r in rows:
        r.insert(desc_ndx, _describe_error(r[desc_ndx - 1] if desc_ndx - 1 < len(r) else ''))

    title = 'Error & Incident Report'
    if fmt == 'excel':
        wb = openpyxl.Workbook(); wb.remove(wb.active)
        xl_build_sheet(wb, 'Errors', title, days, atm, headers, rows)
        log_action('EXPORT', f'Error report (excel) - days={days}, atm={atm}')
        return xl_send(wb, 'Errors')
    elif fmt == 'pdf':
        total_occ = sum(r[4] for r in rows)
        unique_codes = len({r[2] for r in rows})
        unique_atms = len({r[0] for r in rows})
        code_counts = Counter(r[2] for r in rows)
        most_common = code_counts.most_common(1)[0][0] if code_counts else 'N/A'
        kpis = [
            ('Total Errors', f'{total_occ:,}'),
            ('Unique Error Codes', str(unique_codes)),
            ('ATMs Affected', str(unique_atms)),
            ('Most Common', most_common),
        ]
        log_action('EXPORT', f'Error report (pdf) - days={days}, atm={atm}')
        return pdf_send(title, headers, rows, days, atm, 'Errors', kpis=kpis)
    else:
        log_action('EXPORT', f'Error report (csv) - days={days}, atm={atm}')
        return csv_send(headers, rows, 'Errors', title=title, days=days, atm=atm)


# ─── ATM PERFORMANCE ────────────────────────────────────────────────────────────

@bp.route('/report/performance/<fmt>')
@login_required
@role_required(ROLE_VIEWER, ROLE_OPERATOR, ROLE_ADMIN)
def report_performance(fmt):
    days = int(request.args.get('days', 7))

    with get_db() as conn:
        cur = conn.cursor()
        sql = """
            WITH stats AS (
                SELECT atm_id, branch,
                    COUNT(*) AS total,
                    ROUND(100.0*SUM(CASE WHEN status='APPROVED' THEN 1 ELSE 0 END)
                        /NULLIF(COUNT(*),0),2) AS rate,
                    ROUND(COUNT(*)/%s::numeric,1) AS daily
                FROM atm_transactions
                WHERE recorded_at >= NOW() - INTERVAL %s
                GROUP BY atm_id, branch
            ),
            peaks AS (
                SELECT DISTINCT ON (atm_id) atm_id,
                    EXTRACT(HOUR FROM recorded_at) AS hr, COUNT(*) AS cnt
                FROM atm_transactions
                WHERE recorded_at >= NOW() - INTERVAL %s
                GROUP BY atm_id, EXTRACT(HOUR FROM recorded_at)
                ORDER BY atm_id, cnt DESC
            )
            SELECT s.atm_id AS "ATM", s.branch AS "Branch",
                s.total AS "Total Transactions",
                s.rate  AS "Success Rate",
                s.daily AS "Avg Daily Txns",
                CONCAT(p.hr::int,':00-',(p.hr::int+1),':00') AS "Peak Hour",
                RANK() OVER (ORDER BY s.total DESC) AS "Rank"
            FROM stats s JOIN peaks p ON s.atm_id=p.atm_id
            ORDER BY s.total DESC
        """
        days_interval = f"{days} days"
        cur.execute(sql, (days, days_interval, days_interval))
        rows = cur.fetchall(); headers = [d[0] for d in cur.description]
        cur.close()

    title = 'ATM Performance Report'
    if fmt == 'excel':
        wb = openpyxl.Workbook(); wb.remove(wb.active)
        xl_build_sheet(wb, 'Performance', title, days, 'all', headers, rows)
        log_action('EXPORT', f'Performance report (excel) - days={days}')
        return xl_send(wb, 'Performance')
    elif fmt == 'pdf':
        total_txns = sum(r[2] for r in rows)
        avg_rate = round(sum(r[3] for r in rows) / len(rows), 1) if rows else 0
        top_atm = rows[0][0] if rows else 'N/A'
        atm_count = len(rows)
        kpis = [
            ('Total Transactions', f'{total_txns:,}'),
            ('Avg Success Rate', f'{avg_rate}%'),
            ('ATMs Ranked', str(atm_count)),
            ('Top Performer', str(top_atm)),
        ]
        log_action('EXPORT', f'Performance report (pdf) - days={days}')
        return pdf_send(title, headers, rows, days, 'all', 'Performance', kpis=kpis)
    else:
        log_action('EXPORT', f'Performance report (csv) - days={days}')
        return csv_send(headers, rows, 'Performance', title=title, days=days, atm='all')


# ─── ATM AVAILABILITY ───────────────────────────────────────────────────────────

@bp.route('/report/availability/<fmt>')
@login_required
@role_required(ROLE_VIEWER, ROLE_OPERATOR, ROLE_ADMIN)
def report_availability(fmt):
    days = int(request.args.get('days', 7))
    atm  = request.args.get('atm', 'all')

    with get_db() as conn:
        cur = conn.cursor()
        sql = """
            WITH hourly AS (
                SELECT
                    t.atm_id,
                    l.branch,
                    DATE_TRUNC('hour', recorded_at) as hour,
                    COUNT(*) as txn_count
                FROM atm_transactions t
                JOIN atm_locations l ON t.atm_id = l.atm_id
                WHERE recorded_at >= NOW() - INTERVAL %s
            """ + (" AND t.atm_id = %s" if atm != 'all' else "") + """
                GROUP BY t.atm_id, l.branch, DATE_TRUNC('hour', recorded_at)
            ),
            total_hours AS (
                SELECT %s * 24 as expected_hours
            ),
            uptime AS (
                SELECT
                    atm_id,
                    branch,
                    COUNT(*) as active_hours,
                    (SELECT expected_hours FROM total_hours) as expected_hours,
                    ROUND(100.0 * COUNT(*) / (SELECT expected_hours FROM total_hours), 2) as uptime_pct
                FROM hourly
                GROUP BY atm_id, branch
            )
            SELECT
                atm_id as "ATM ID",
                branch as "Branch",
                active_hours as "Active Hours",
                expected_hours as "Expected Hours",
                (expected_hours - active_hours) as "Downtime Hours",
                uptime_pct as "Uptime",
                CASE
                    WHEN uptime_pct >= 99.9 THEN 'Excellent'
                    WHEN uptime_pct >= 99.0 THEN 'Good'
                    WHEN uptime_pct >= 95.0 THEN 'Acceptable'
                    ELSE 'Below Target'
                END as "SLA Status"
            FROM uptime
            ORDER BY uptime_pct DESC
        """
        params = [f"{days} days"]
        if atm != 'all':
            params.append(atm)
        params.append(days)
        cur.execute(sql, tuple(params))
        rows = cur.fetchall()
        headers = [d[0] for d in cur.description]
        cur.close()

    title = 'ATM Availability Report'
    if fmt == 'excel':
        wb = openpyxl.Workbook(); wb.remove(wb.active)
        ws = wb.create_sheet('Availability')
        xl_header(ws, title, days, atm)
        ws.append(headers)
        xl_style_row(ws, ws.max_row, len(headers))
        data_start = ws.max_row + 1
        for row in rows:
            ws.append(list(row))
        xl_style_data_rows(ws, data_start, len(headers))
        colors_map = {
            'Excellent': 'C6EFCE',
            'Good': 'C6EFCE',
            'Acceptable': 'FFEB9C',
            'Below Target': 'FFC7CE'
        }
        for idx, row in enumerate(rows):
            sla_val = str(row[-1] or '') if row else ''
            fill_color = colors_map.get(sla_val, 'FFFFFF')
            cell = ws.cell(row=data_start + idx, column=len(headers))
            cell.fill = openpyxl.styles.PatternFill('solid', fgColor=fill_color)
        xl_autosize(ws)
        log_action('EXPORT', f'Availability report (excel) - days={days}, atm={atm}')
        return xl_send(wb, 'Availability')
    elif fmt == 'pdf':
        avg_uptime = round(sum(r[5] for r in rows) / len(rows), 2) if rows else 0
        excellent = sum(1 for r in rows if r[6] == 'Excellent')
        below_target = sum(1 for r in rows if r[6] == 'Below Target')
        kpis = [
            ('Avg Uptime', f'{avg_uptime}%'),
            ('Excellent SLA', str(excellent)),
            ('Below Target', str(below_target)),
            ('ATMs Monitored', str(len(rows))),
        ]
        log_action('EXPORT', f'Availability report (pdf) - days={days}, atm={atm}')
        return pdf_send(title, headers, rows, days, atm, 'Availability', kpis=kpis)
    else:
        log_action('EXPORT', f'Availability report (csv) - days={days}, atm={atm}')
        return csv_send(headers, rows, 'Availability', title=title, days=days, atm=atm)


# ─── COMPLETE MANAGEMENT REPORT ─────────────────────────────────────────────────

@bp.route('/report/full/<fmt>')
@login_required
@role_required(ROLE_VIEWER, ROLE_OPERATOR, ROLE_ADMIN)
def report_full(fmt):
    days = int(request.args.get('days', 7))
    atm  = request.args.get('atm', 'all')
    interval_str = f"{days} days"
    with get_db() as conn:

        if fmt == 'excel':
            wb = openpyxl.Workbook(); ws = wb.active
            ws.title = 'Complete Report'
            xl_header(ws, 'Complete Management Report', days, atm)

            # 1. Transactions
            cur = conn.cursor()
            sql1 = """
                SELECT atm_id, branch, COUNT(*),
                    SUM(CASE WHEN status='APPROVED' THEN 1 ELSE 0 END),
                    SUM(CASE WHEN status='DECLINED' THEN 1 ELSE 0 END),
                    SUM(CASE WHEN status='ERROR'    THEN 1 ELSE 0 END),
                    ROUND(100.0*SUM(CASE WHEN status='APPROVED' THEN 1 ELSE 0 END)/NULLIF(COUNT(*),0),2),
                    COALESCE(SUM(CASE WHEN status='APPROVED' AND txn_type='WITHDRAWAL' THEN amount ELSE 0 END),0)
                FROM atm_transactions WHERE recorded_at >= NOW() - INTERVAL %s
            """
            p1 = [interval_str]
            if atm != 'all': sql1 += " AND atm_id = %s"; p1.append(atm)
            sql1 += " GROUP BY atm_id, branch ORDER BY atm_id"
            cur.execute(sql1, p1)
            h1 = ['ATM', 'Branch', 'Total Txns', 'Approved', 'Declined', 'Errors', 'Success Rate %', 'Cash Dispensed ETB']
            xl_append_section(ws, '1. Transaction Summary', h1, cur.fetchall())
            cur.close()

            # 2. Cash
            cur = conn.cursor()
            sql2 = """
                SELECT atm_id, branch, COALESCE(SUM(amount),0),
                    COALESCE(ROUND(AVG(amount),2),0), COUNT(*), COALESCE(MAX(amount),0)
                FROM atm_transactions WHERE status='APPROVED' AND txn_type='WITHDRAWAL'
                AND recorded_at >= NOW() - INTERVAL %s
            """
            p2 = [interval_str]
            if atm != 'all': sql2 += " AND atm_id = %s"; p2.append(atm)
            sql2 += " GROUP BY atm_id, branch ORDER BY 3 DESC"
            cur.execute(sql2, p2)
            h2 = ['ATM', 'Branch', 'Total Dispensed ETB', 'Avg Withdrawal ETB', 'Withdrawal Count', 'Largest Withdrawal ETB']
            xl_append_section(ws, '2. Cash Level & Dispensing', h2, cur.fetchall())
            cur.close()

            # 3. Errors
            cur = conn.cursor()
            sql3 = """
                SELECT atm_id, branch, error_code, COUNT(*),
                    TO_CHAR(MIN(recorded_at) AT TIME ZONE 'Africa/Addis_Ababa','YYYY-MM-DD HH24:MI'),
                    TO_CHAR(MAX(recorded_at) AT TIME ZONE 'Africa/Addis_Ababa','YYYY-MM-DD HH24:MI')
                FROM atm_transactions WHERE status='ERROR' AND error_code IS NOT NULL
                AND recorded_at >= NOW() - INTERVAL %s
            """
            p3 = [interval_str]
            if atm != 'all': sql3 += " AND atm_id = %s"; p3.append(atm)
            sql3 += " GROUP BY atm_id, branch, error_code ORDER BY 4 DESC"
            cur.execute(sql3, p3)
            rows3 = cur.fetchall()
            h3 = ['ATM', 'Branch', 'Error Code', 'Description', 'Occurrences', 'First Seen', 'Last Seen']
            rows3_desc = [list(r) for r in rows3]
            for r in rows3_desc:
                r.insert(3, _describe_error(r[2]))
            xl_append_section(ws, '3. Error & Incident Log', h3, rows3_desc)
            cur.close()

            # 4. Performance
            cur = conn.cursor()
            sql4 = """
                WITH stats AS (
                    SELECT atm_id, branch, COUNT(*) AS total,
                        ROUND(100.0*SUM(CASE WHEN status='APPROVED' THEN 1 ELSE 0 END)/NULLIF(COUNT(*),0),2) AS rate,
                        ROUND(COUNT(*)/%s::numeric,1) AS daily
                    FROM atm_transactions WHERE recorded_at >= NOW() - INTERVAL %s
                    GROUP BY atm_id, branch
                ),
                peaks AS (
                    SELECT DISTINCT ON (atm_id) atm_id,
                        EXTRACT(HOUR FROM recorded_at) AS hr, COUNT(*) AS cnt
                    FROM atm_transactions WHERE recorded_at >= NOW() - INTERVAL %s
                    GROUP BY atm_id, EXTRACT(HOUR FROM recorded_at)
                    ORDER BY atm_id, cnt DESC
                )
                SELECT s.atm_id, s.branch, s.total, s.rate, s.daily,
                    CONCAT(p.hr::int,':00-',(p.hr::int+1),':00'),
                    RANK() OVER (ORDER BY s.total DESC)
                FROM stats s JOIN peaks p ON s.atm_id=p.atm_id ORDER BY s.total DESC
            """
            cur.execute(sql4, (days, interval_str, interval_str))
            h4 = ['ATM', 'Branch', 'Total Transactions', 'Success Rate %', 'Avg Daily Txns', 'Peak Hour', 'Rank']
            xl_append_section(ws, '4. ATM Performance Metrics', h4, cur.fetchall())
            cur.close()

            xl_autosize(ws)

            log_action('EXPORT', f'Complete report (excel) - days={days}, atm={atm}')
            return xl_send(wb, 'Complete_Report')

        elif fmt == 'csv':
            buf = io.StringIO()
            w = csv.writer(buf)
            w.writerow(['DASHEN BANK S.C. — ATM MONITORING SYSTEM'])
            w.writerow(['COMPLETE MANAGEMENT REPORT'])
            w.writerow([f'Period: Last {days} days', f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} EAT'])
            w.writerow([])

            sections = [
                ('1. TRANSACTION SUMMARY',
                 "SELECT atm_id, branch, COUNT(*), SUM(CASE WHEN status='APPROVED' THEN 1 ELSE 0 END), SUM(CASE WHEN status='DECLINED' THEN 1 ELSE 0 END), SUM(CASE WHEN status='ERROR' THEN 1 ELSE 0 END), ROUND(100.0*SUM(CASE WHEN status='APPROVED' THEN 1 ELSE 0 END)/NULLIF(COUNT(*),0),2), COALESCE(SUM(CASE WHEN status='APPROVED' AND txn_type='WITHDRAWAL' THEN amount ELSE 0 END),0) FROM atm_transactions WHERE recorded_at >= NOW() - INTERVAL %s GROUP BY atm_id, branch ORDER BY atm_id",
                 [interval_str],
                 ['ATM', 'Branch', 'Total Txns', 'Approved', 'Declined', 'Errors', 'Success Rate %', 'Cash Dispensed ETB']),
                ('2. CASH LEVEL & DISPENSING',
                 "SELECT atm_id, branch, COALESCE(SUM(amount),0), COALESCE(ROUND(AVG(amount),2),0), COUNT(*), COALESCE(MAX(amount),0) FROM atm_transactions WHERE status='APPROVED' AND txn_type='WITHDRAWAL' AND recorded_at >= NOW() - INTERVAL %s GROUP BY atm_id, branch ORDER BY 3 DESC",
                 [interval_str],
                 ['ATM', 'Branch', 'Total Dispensed ETB', 'Avg Withdrawal ETB', 'Withdrawal Count', 'Largest Withdrawal ETB']),
                ('3. ERROR & INCIDENT LOG',
                 "SELECT atm_id, branch, error_code, COUNT(*), TO_CHAR(MIN(recorded_at) AT TIME ZONE 'Africa/Addis_Ababa','YYYY-MM-DD HH24:MI'), TO_CHAR(MAX(recorded_at) AT TIME ZONE 'Africa/Addis_Ababa','YYYY-MM-DD HH24:MI') FROM atm_transactions WHERE status='ERROR' AND error_code IS NOT NULL AND recorded_at >= NOW() - INTERVAL %s GROUP BY atm_id, branch, error_code ORDER BY 4 DESC",
                 [interval_str],
                 ['ATM', 'Branch', 'Error Code', 'Occurrences', 'First Seen', 'Last Seen']),
            ]

            for sec_title, sql, params, headers in sections:
                cur = conn.cursor()
                cur.execute(sql, params)
                rows = cur.fetchall()
                cur.close()
                if 'ERROR' in sec_title:
                    headers = headers[:3] + ['Description'] + headers[3:]
                    rows = [list(r) for r in rows]
                    for r in rows:
                        r.insert(3, _describe_error(r[2]))
                w.writerow([sec_title])
                w.writerow(headers)
                w.writerows([[str(v) if v is not None else '—' for v in r] for r in rows])
                w.writerow([])

            cur = conn.cursor()
            sql4 = """
                WITH stats AS (
                    SELECT atm_id, branch, COUNT(*) AS total,
                        ROUND(100.0*SUM(CASE WHEN status='APPROVED' THEN 1 ELSE 0 END)/NULLIF(COUNT(*),0),2) AS rate,
                        ROUND(COUNT(*)/%s::numeric,1) AS daily
                    FROM atm_transactions WHERE recorded_at >= NOW() - INTERVAL %s GROUP BY atm_id, branch
                ),
                peaks AS (
                    SELECT DISTINCT ON (atm_id) atm_id, EXTRACT(HOUR FROM recorded_at) AS hr, COUNT(*) AS cnt
                    FROM atm_transactions WHERE recorded_at >= NOW() - INTERVAL %s
                    GROUP BY atm_id, EXTRACT(HOUR FROM recorded_at) ORDER BY atm_id, cnt DESC
                )
                SELECT s.atm_id, s.branch, s.total, s.rate, s.daily,
                    CONCAT(p.hr::int,':00-',(p.hr::int+1),':00'), RANK() OVER (ORDER BY s.total DESC)
                FROM stats s JOIN peaks p ON s.atm_id=p.atm_id ORDER BY s.total DESC
            """
            cur.execute(sql4, (days, interval_str, interval_str))
            rows4 = cur.fetchall(); cur.close()
            w.writerow(['4. ATM PERFORMANCE METRICS'])
            w.writerow(['ATM', 'Branch', 'Total Transactions', 'Success Rate %', 'Avg Daily Txns', 'Peak Hour', 'Rank'])
            w.writerows([[str(v) if v is not None else '—' for v in r] for r in rows4])
            w.writerow([])
            w.writerow(['Dashen Bank ATM Monitoring System | Confidential Management Report'])

            log_action('EXPORT', f'Complete report (csv) - days={days}, atm={atm}')
            buf.seek(0)
            return send_file(
                io.BytesIO(buf.getvalue().encode('utf-8-sig')),
                mimetype='text/csv', as_attachment=True,
                download_name=f'Dashen_ATM_Complete_Report_{datetime.now().strftime("%Y%m%d")}.csv'
            )

        else:
            buf = io.BytesIO()
            doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                    rightMargin=1.5 * cm, leftMargin=1.5 * cm,
                                    topMargin=2 * cm, bottomMargin=1.5 * cm)
            style_bank = ParagraphStyle('_h1', fontSize=18, leading=22, alignment=TA_CENTER, spaceAfter=4)
            style_sys  = ParagraphStyle('_h2', fontSize=11, leading=14, alignment=TA_CENTER, spaceAfter=8)
            style_full = ParagraphStyle('_h3', fontSize=14, leading=18, alignment=TA_CENTER, spaceAfter=8)
            style_meta = ParagraphStyle('_h4', fontSize=9, leading=12, alignment=TA_CENTER)
            style_sec  = ParagraphStyle('_h_sec', fontSize=12, leading=15, spaceBefore=12, spaceAfter=8)

            story = []
            if os.path.exists(LOGO_PATH):
                img = Image(LOGO_PATH, width=6 * cm, height=1.8 * cm, kind='proportional')
                img.hAlign = 'CENTER'
                story.append(img)
                story.append(Spacer(1, 0.5 * cm))

            story += [
                Paragraph('<b><font color="#012169">DASHEN BANK S.C.</font></b>', style_bank),
                Paragraph('<font color="#273274">ATM MONITORING SYSTEM</font>', style_sys),
                HRFlowable(width='60%', thickness=2.5, color=colors.HexColor('#FDD79A'), spaceAfter=12),
                Paragraph('<b><font color="#273274">COMPLETE MANAGEMENT REPORT</font></b>', style_full),
                Paragraph(f'<font color="#64748B">Period: Last {days} days  |  Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} EAT</font>', style_meta),
                Spacer(1, 1.2 * cm)
            ]

            cur = conn.cursor()

            # 1. Transaction Summary
            s1_sql = "SELECT atm_id, branch, COUNT(*), SUM(CASE WHEN status='APPROVED' THEN 1 ELSE 0 END), SUM(CASE WHEN status='DECLINED' THEN 1 ELSE 0 END), SUM(CASE WHEN status='ERROR' THEN 1 ELSE 0 END), ROUND(100.0*SUM(CASE WHEN status='APPROVED' THEN 1 ELSE 0 END)/NULLIF(COUNT(*),0),2), COALESCE(SUM(CASE WHEN status='APPROVED' AND txn_type='WITHDRAWAL' THEN amount ELSE 0 END),0) FROM atm_transactions WHERE recorded_at >= NOW() - INTERVAL %s GROUP BY atm_id, branch ORDER BY atm_id"
            cur.execute(s1_sql, [interval_str])
            story += [Paragraph('<b><font color="#012169">1. Transaction Summary</font></b>', style_sec),
                      mktable(['ATM', 'Branch', 'Total Txns', 'Approved', 'Declined', 'Errors', 'Success Rate %', 'Cash Dispensed ETB'], cur.fetchall()),
                      Spacer(1, 0.8 * cm)]

            # 2. Cash Level
            s2_sql = "SELECT atm_id, branch, COALESCE(SUM(amount),0), COALESCE(ROUND(AVG(amount),2),0), COUNT(*), COALESCE(MAX(amount),0) FROM atm_transactions WHERE status='APPROVED' AND txn_type='WITHDRAWAL' AND recorded_at >= NOW() - INTERVAL %s GROUP BY atm_id, branch ORDER BY 3 DESC"
            cur.execute(s2_sql, [interval_str])
            story += [Paragraph('<b><font color="#012169">2. Cash Level & Dispensing</font></b>', style_sec),
                      mktable(['ATM', 'Branch', 'Total Dispensed ETB', 'Avg Withdrawal ETB', 'Withdrawal Count', 'Largest Withdrawal ETB'], cur.fetchall()),
                      Spacer(1, 0.8 * cm)]

            # 3. Error & Incident
            s3_sql = "SELECT atm_id, branch, error_code, COUNT(*), TO_CHAR(MIN(recorded_at) AT TIME ZONE 'Africa/Addis_Ababa','YYYY-MM-DD HH24:MI'), TO_CHAR(MAX(recorded_at) AT TIME ZONE 'Africa/Addis_Ababa','YYYY-MM-DD HH24:MI') FROM atm_transactions WHERE status='ERROR' AND error_code IS NOT NULL AND recorded_at >= NOW() - INTERVAL %s GROUP BY atm_id, branch, error_code ORDER BY 4 DESC"
            cur.execute(s3_sql, [interval_str])
            res3 = cur.fetchall()
            if res3:
                res3_desc = [list(r) for r in res3]
                for r in res3_desc:
                    r.insert(3, _describe_error(r[2]))
                story += [Paragraph('<b><font color="#012169">3. Error & Incident Log</font></b>', style_sec),
                          mktable(['ATM', 'Branch', 'Error Code', 'Description', 'Occurrences', 'First Seen', 'Last Seen'], res3_desc),
                          Spacer(1, 0.8 * cm)]

            # 4. Performance Metrics
            s4_sql = """
                WITH stats AS (
                    SELECT atm_id, branch, COUNT(*) AS total,
                        ROUND(100.0*SUM(CASE WHEN status='APPROVED' THEN 1 ELSE 0 END)/NULLIF(COUNT(*),0),2) AS rate,
                        ROUND(COUNT(*)/%s::numeric,1) AS daily
                    FROM atm_transactions WHERE recorded_at >= NOW() - INTERVAL %s GROUP BY atm_id, branch
                ),
                peaks AS (
                    SELECT DISTINCT ON (atm_id) atm_id, EXTRACT(HOUR FROM recorded_at) AS hr, COUNT(*) AS cnt
                    FROM atm_transactions WHERE recorded_at >= NOW() - INTERVAL %s
                    GROUP BY atm_id, EXTRACT(HOUR FROM recorded_at) ORDER BY atm_id, cnt DESC
                )
                SELECT s.atm_id, s.branch, s.total, s.rate, s.daily,
                    CONCAT(p.hr::int,':00-',(p.hr::int+1),':00'), RANK() OVER (ORDER BY s.total DESC)
                FROM stats s JOIN peaks p ON s.atm_id=p.atm_id ORDER BY s.total DESC
            """
            cur.execute(s4_sql, (days, interval_str, interval_str))
            res4 = cur.fetchall()
            if res4:
                story += [PageBreak(),
                          Paragraph('<b><font color="#012169">4. ATM Performance Metrics</font></b>', style_sec),
                          mktable(['ATM', 'Branch', 'Total Transactions', 'Success Rate %', 'Avg Daily Txns', 'Peak Hour', 'Rank'], res4),
                          Spacer(1, 0.8 * cm)]

            story.append(Paragraph(
                f'<font size="8" color="#64748B">Dashen Bank ATM Monitoring System  |  Confidential Management Report  |  {datetime.now().strftime("%Y-%m-%d")}</font>',
                style_meta
            ))

            cur.close()
            log_action('EXPORT', f'Complete report (pdf) - days={days}, atm={atm}')
            doc.build(story); buf.seek(0)
            return send_file(buf, mimetype='application/pdf', as_attachment=True,
                             download_name=f'Dashen_ATM_Complete_Report_{datetime.now().strftime("%Y%m%d")}.pdf')


# ─── HARDWARE INVENTORY VIEW (NetXMS equivalent) ────────────────────────────────

@bp.route('/inventory')
@login_required
@role_required(ROLE_VIEWER, ROLE_OPERATOR, ROLE_ADMIN)
def inventory():
    try:
        with get_db() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT
                    l.atm_name,
                    l.branch,
                    l.atm_id,
                    l.vendor,
                    l.model,
                    l.city,
                    l.district,
                    l.install_date,
                    s.state,
                    s.last_seen
                FROM atm_locations l
                LEFT JOIN atm_current_state s ON l.atm_id = s.atm_id
                ORDER BY l.vendor, l.atm_id
            """)
            atms = cur.fetchall()
            cur.close()

        rows_html = ''.join([f'''
            <tr>
                <td><b>{r[0]}</b></td>
                <td>{r[1]}</td>
                <td style="font-family:monospace">{r[2]}</td>
                <td><span style="background:{'#cce5ff' if r[3]=='NCR' else '#e2d9f3'};
                    padding:2px 8px;border-radius:10px;font-size:11px">{r[3]}</span></td>
                <td>{r[4] or '—'}</td>
                <td>{r[5]}</td>
                <td>{r[6]}</td>
                <td>{r[7] or '—'}</td>
                <td><span style="color:{'green' if r[8]=='IN_SERVICE' else 'red'};
                    font-weight:600">{r[8] or 'UNKNOWN'}</span></td>
            </tr>
        ''' for r in atms])

        return f'''<!DOCTYPE html>
<html><head><title>ATM Inventory — Dashen Bank</title>
<style>
body{{font-family:Segoe UI,sans-serif;background:#f0f4f8;}}
.hdr{{background:linear-gradient(135deg,#1B3A6B,#2d5a9e);color:white;padding:16px 32px;}}
.gold{{height:4px;background:#C9A84C;}}
.container{{max-width:1400px;margin:24px auto;padding:0 20px;}}
table{{width:100%;border-collapse:collapse;background:white;border-radius:10px;
    box-shadow:0 2px 8px rgba(0,0,0,0.08);overflow:hidden;}}
th{{background:#1B3A6B;color:white;padding:10px 12px;text-align:left;font-size:11px;}}
td{{padding:9px 12px;border-bottom:1px solid #f0f0f0;font-size:12px;}}
tr:hover td{{background:#f8f9ff;}}
</style></head>
<body>
<div class="hdr"><h1>ATM Hardware Inventory</h1>
<p>All ATMs — NCR and GRG</p></div>
<div class="gold"></div>
<div class="container">
<p style="margin-bottom:12px;color:#666;font-size:13px">
    Total: {len(atms)} ATMs |
    NCR: {sum(1 for r in atms if r[3]=='NCR')} |
    GRG: {sum(1 for r in atms if r[3]=='GRG')}
</p>
<table>
<thead><tr>
    <th>Name</th><th>Branch</th><th>Terminal ID</th>
    <th>Vendor</th><th>Model</th><th>City</th>
    <th>District</th><th>Install Date</th><th>State</th>
</tr></thead>
<tbody>{rows_html}</tbody>
</table></div></body></html>'''
    except Exception as e:
        return f'<h1>Error</h1><pre>{e}</pre>', 500

